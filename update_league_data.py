#!/usr/bin/env python3
"""
Fantasy Golf League Data Extractor
===================================
This script reads your FG_2026_AI_Scores.xlsx file and generates league-data.json
for the Fantasy Golf web app.

USAGE:
1. Place this script in the same folder as FG_2026_AI_Scores.xlsx
2. Double-click the script (or run: python update_league_data.py)
3. Upload the generated league-data.json file alongside your web app
"""

import openpyxl
import json
from datetime import datetime

def extract_league_data(excel_file='FG_2026_AI_Scores.xlsx'):
    """Extract all league data from Excel file."""
    
    print(f"Reading {excel_file}...")
    wb = openpyxl.load_workbook(excel_file, data_only=True)
    
    # Extract team data from 1st Half sheet
    teams_data = extract_teams(wb)
    
    # Extract transactions from Drop-Add Log
    transactions = extract_transactions(wb)
    
    # Find available players (players not on any roster)
    available_players = find_available_players(teams_data)
    
    league_data = {
        'teams': teams_data,
        'transactions': transactions,
        'availablePlayers': sorted(available_players),
        'lastUpdated': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    }
    
    return league_data


def extract_teams(wb):
    """Extract all team rosters and points from the 1st Half sheet."""
    ws = wb['1st Half']
    teams = {}
    
    # Team rows start at these positions
    team_start_rows = {
        'THE ARMY': 3,
        'PETE/SWAN': 13,
        'BOO/SMITTY': 23,
        'MASON/WRIGHT': 33,
        'SLAYDEN': 43,
        'JR/EGAN': 53,
        'ZY': 63,
        'JC/WES': 73,
        'BILLY/STICK': 83,
        'CARSON': 93
    }
    
    for team_name, start_row in team_start_rows.items():
        players = []
        
        # Read 8 players (rows after team name)
        for offset in range(1, 9):
            player_row = start_row + offset
            player_name = ws.cell(row=player_row, column=1).value
            if player_name and player_name != 'TOTAL':
                players.append(player_name)
        
        # Get team total points (TOTAL row, column S which is 19)
        total_row = start_row + 9  # TOTAL is always 9 rows after team name
        total_points = ws.cell(row=total_row, column=19).value or 0
        
        # Count moves used from Drop-Add Log
        moves_used = count_team_moves(wb, team_name)
        
        teams[team_name] = {
            'players': players,
            'points': float(total_points),
            'firstHalf': float(total_points),  # Since we're in season, all points are 1st half for now
            'secondHalf': 0,
            'movesUsed': moves_used
        }
    
    print(f"✓ Extracted {len(teams)} teams")
    return teams


def count_team_moves(wb, team_name):
    """Count how many drop/add moves a team has used."""
    ws = wb['Drop-Add Log']
    moves = 0
    
    # Start from row 3 (first data row after headers)
    for row in range(3, 100):
        team_cell = ws.cell(row=row, column=2).value
        if team_cell == team_name:
            moves += 1
    
    return moves


def extract_transactions(wb):
    """Extract all drop/add transactions from the log."""
    ws = wb['Drop-Add Log']
    transactions = []
    
    # Start from row 3 (first data row)
    for row in range(3, 100):
        date_val = ws.cell(row=row, column=1).value
        team_val = ws.cell(row=row, column=2).value
        dropped = ws.cell(row=row, column=3).value
        added = ws.cell(row=row, column=4).value
        
        if not team_val:  # Stop at empty rows
            break
            
        # Format date
        if isinstance(date_val, datetime):
            date_str = date_val.strftime('%m/%d/%Y')
        else:
            date_str = str(date_val) if date_val else ''
        
        # Add DROP transaction
        if dropped:
            transactions.append({
                'team': team_val,
                'action': 'DROP',
                'player': dropped,
                'date': date_str
            })
        
        # Add ADD transaction
        if added:
            transactions.append({
                'team': team_val,
                'action': 'ADD',
                'player': added,
                'date': date_str
            })
    
    print(f"✓ Extracted {len(transactions)} transactions")
    return transactions


def find_available_players(teams_data):
    """Find players not currently on any roster."""
    
    # All PGA Tour players (you can expand this list as needed)
    all_players = {
        'Rickie Fowler', 'Cameron Smith', 'Justin Lower', 'Patrick Reed',
        'Billy Gotterup', 'Webb Simpson', 'Gary Woodland', 'Lucas Glover',
        'Lee Hodges', 'Taylor Moore', 'Kevin Yu', 'Zac Blair',
        'Sami Valimaki', 'Marco Penge', 'Charley Hoffman', 'Luke List',
        'Adam Hadwin', 'Doug Ghim', 'Cameron Champ', 'Mackenzie Hughes'
    }
    
    # Get all rostered players
    rostered = set()
    for team_data in teams_data.values():
        rostered.update(team_data['players'])
    
    # Available = all players minus rostered
    available = all_players - rostered
    
    print(f"✓ Found {len(available)} available players")
    return list(available)


def save_json(data, filename='league-data.json'):
    """Save league data to JSON file."""
    with open(filename, 'w') as f:
        json.dump(data, f, indent=2)
    print(f"✓ Saved to {filename}")


if __name__ == '__main__':
    try:
        print("\n" + "="*60)
        print("FANTASY GOLF - League Data Extractor")
        print("="*60 + "\n")
        
        league_data = extract_league_data()
        save_json(league_data)
        
        print("\n" + "="*60)
        print("SUCCESS! Your league-data.json file is ready.")
        print("="*60)
        print("\nNEXT STEPS:")
        print("1. Upload league-data.json to the same location as your app")
        print("2. Refresh your Fantasy Golf app in your browser")
        print("3. The app will now show your updated data!")
        print("\n")
        
    except FileNotFoundError:
        print("\n❌ ERROR: Could not find FG_2026_AI_Scores.xlsx")
        print("   Make sure this script is in the same folder as your Excel file.\n")
    except Exception as e:
        print(f"\n❌ ERROR: {e}\n")
